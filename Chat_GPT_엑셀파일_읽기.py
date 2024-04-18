from openpyxl import load_workbook

# 엑셀 파일 경로
file_path = r'c:\work\products.xlsx'

# 엑셀 파일 로드
workbook = load_workbook(filename=file_path)

# 첫 번째 시트 선택
sheet = workbook.active

# 시트의 각 행을 순회하며 출력
for row in sheet.iter_rows(values_only=True):
    print(row)

# 작업 완료 후 파일 닫기
workbook.close()
